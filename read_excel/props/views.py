# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from django.http import HttpResponse
import requests
import json

from django.shortcuts import render
from django.conf import settings

import openpyxl
from openpyxl.writer.excel import save_virtual_workbook

API_KEY = "AIzaSyCqjCZP5ljNyRYyLXv-BfJrP6pkHb2YoAQ"

def index(request):
    return HttpResponse("Hello, world. You're at props")

def get_location(address):
    add = "+".join(address.split(" "))
    response = requests.get('https://maps.googleapis.com/maps/api/geocode/json?address=' + add + '&key='+API_KEY)
    json_response = json.loads(response.text)
    if json_response["status"] == "OK":
        location = json_response["result"][0]["geometry"]["location"]
    else:
        location = {
                    "lat": "000",
                    "lng": "-000",
                    }
    return location


def read_file(myfile):
    wb = openpyxl.load_workbook(myfile)
    worksheet = wb["Sheet1"]
    excel_data = list()
    row_count = 0
    for row in worksheet.iter_rows():
        row_data = {}
        if row_count != 0:
            for cell in row:
                location = get_location(str(cell.value))
                row_data["address"] = str(cell.value)
                row_data["lat"] = location["lat"]
                row_data["lng"] = location["lng"]
                break
            excel_data.append(row_data)
        row_count = row_count + 1

    return excel_data

def write_file(data, file_name):
    wb = openpyxl.load_workbook(filename=file_name)
    ws = wb.worksheets[0]
    for index, v in enumerate(data):
        print (v)
        ws.cell(row=(index+2), column=1).value = v["address"]
        ws.cell(row=(index+2), column=2).value = v["lat"]
        ws.cell(row=(index+2), column=3).value = v["lng"]

    response = HttpResponse(content_type = 'application/vnd.openzmlformat-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'
    wb.save(response)
    return response

def upload(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        data = read_file(myfile)
        response = write_file(data, myfile.name)

        return response
    return render(request, 'props/index.html')
